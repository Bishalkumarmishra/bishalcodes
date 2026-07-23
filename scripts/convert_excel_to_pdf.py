import sys
import os
import subprocess
import shutil

def optimize_excel_layout(input_path, output_path):
    # Only support XLSX files for openpyxl optimization
    if not input_path.lower().endswith('.xlsx'):
        shutil.copy(input_path, output_path)
        return False
        
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        print("Optimizing Excel sheet layout and gridlines with openpyxl...")
        wb = openpyxl.load_workbook(input_path)
        
        for ws in wb.worksheets:
            # 1. Enable gridlines visibility in view and print
            if ws.views.sheetView:
                ws.views.sheetView[0].showGridLines = True
            else:
                ws.sheet_view.showGridLines = True
                
            ws.print_options.gridLines = True
            
            # 2. Configure page setup scaling: fit all columns wide on 1 page
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0 # 0/False/None lets page height grow naturally
            
            # fitToPage must be set to True for fitToWidth/fitToHeight settings to apply
            if ws.sheet_properties.pageSetUpPr:
                ws.sheet_properties.pageSetUpPr.fitToPage = True
            else:
                from openpyxl.worksheet.properties import PageSetupProperties
                ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
                
            # 3. Auto-fit column widths to prevent text cropping and make spacing tight
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                # Adjust column width with a minimum width of 10 and padding of 3
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
        wb.save(output_path)
        print("Excel layout optimization complete.")
        return True
    except Exception as e:
        print(f"Failed to optimize Excel layout: {e}. Proceeding with original file.")
        shutil.copy(input_path, output_path)
        return False

def convert_excel_to_pdf(excel_path, pdf_path):
    # Try using win32com (Excel COM API) on Windows
    try:
        import win32com.client
        print("Using Excel COM API for conversion...")
        
        # Initialize Excel application
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Absolute paths are required for Excel COM
        abs_excel_path = os.path.abspath(excel_path)
        abs_pdf_path = os.path.abspath(pdf_path)
        
        # Open workbook
        wb = excel.Workbooks.Open(abs_excel_path)
        
        # Enforce page settings on COM object to double-insure gridlines and sheet scaling
        for ws in wb.Worksheets:
            try:
                ws.PageSetup.PrintGridlines = True
                ws.PageSetup.Zoom = False
                ws.PageSetup.FitToPagesWide = 1
                ws.PageSetup.FitToPagesTall = False
            except Exception as ex:
                print(f"COM PageSetup styling ignored: {ex}")
        
        # Export as PDF (type 0 = PDF format)
        wb.ExportAsFixedFormat(0, abs_pdf_path)
        
        wb.Close(False)
        excel.Quit()
        return True
    except Exception as e:
        print(f"Excel COM API conversion failed or not supported: {e}")
        
    # Try using LibreOffice headless conversion
    try:
        print("Trying LibreOffice headless conversion...")
        output_dir = os.path.dirname(pdf_path) or "."
        cmd = ["soffice", "--headless", "--convert-to", "pdf", "--outdir", output_dir, excel_path]
        subprocess.run(cmd, check=True)
        
        # LibreOffice outputs to excel_name.pdf in output_dir, so rename if needed
        filename = os.path.basename(excel_path)
        base_name, _ = os.path.splitext(filename)
        default_output = os.path.join(output_dir, f"{base_name}.pdf")
        
        if os.path.exists(default_output) and os.path.abspath(default_output) != os.path.abspath(pdf_path):
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
            os.rename(default_output, pdf_path)
        return True
    except Exception as e:
        print(f"LibreOffice conversion failed: {e}")
        
    raise Exception("No PDF conversion engine (Microsoft Excel or LibreOffice) could be loaded on the server.")

def main():
    if len(sys.argv) < 3:
        print("Usage: python convert_excel_to_pdf.py <excel_path> <pdf_path>")
        sys.exit(1)
        
    excel_path = sys.argv[1]
    pdf_path = sys.argv[2]
    
    # Path for pre-processed optimized Excel file
    scratch_dir = os.path.join(os.getcwd(), 'scratch')
    os.makedirs(scratch_dir, exist_ok=True)
    temp_excel_path = os.path.join(scratch_dir, f"optimized_{os.path.basename(excel_path)}")
    
    try:
        # Pre-process Excel to set print gridlines, fit columns, and configure scaling
        optimize_excel_layout(excel_path, temp_excel_path)
        
        # Convert optimized workbook
        convert_excel_to_pdf(temp_excel_path, pdf_path)
        print("Conversion completed successfully!")
        
        # Clean up temp file
        try:
            os.remove(temp_excel_path)
        except:
            pass
            
    except Exception as e:
        print(f"Error during conversion: {str(e)}")
        try:
            if os.path.exists(temp_excel_path):
                os.remove(temp_excel_path)
        except:
            pass
        sys.exit(1)

if __name__ == '__main__':
    main()
