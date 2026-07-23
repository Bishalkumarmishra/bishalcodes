import sys
import os
import pdfplumber
import openpyxl

def convert_pdf_to_excel(pdf_path, excel_path):
    print(f"Extracting tables from PDF: {pdf_path}")
    
    # Initialize openpyxl workbook
    wb = openpyxl.Workbook()
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Load PDF
    with pdfplumber.open(pdf_path) as pdf:
        print(f"Total pages in PDF: {len(pdf.pages)}")
        
        sheet_count = 0
        for page_idx, page in enumerate(pdf.pages):
            # Extract tables from the page
            tables = page.extract_tables()
            if not tables:
                # If no tables found, try extracting text and putting it into rows just in case
                text = page.extract_text()
                if text:
                    sheet_count += 1
                    ws = wb.create_sheet(title=f"Page {page_idx + 1}")
                    # Write lines of text as rows
                    for line_idx, line in enumerate(text.split('\n')):
                        ws.cell(row=line_idx + 1, column=1, value=line.strip())
                    ws.views.sheetView[0].showGridLines = True
                continue
                
            for table_idx, table in enumerate(tables):
                sheet_count += 1
                # Worksheet titles cannot exceed 31 characters
                title = f"Sheet{sheet_count}"
                ws = wb.create_sheet(title=title)
                
                # Configure gridlines visibility
                ws.views.sheetView[0].showGridLines = True
                
                # Write table content
                for row_idx, row in enumerate(table):
                    for col_idx, val in enumerate(row):
                        clean_val = str(val).strip() if val is not None else ""
                        ws.cell(row=row_idx + 1, column=col_idx + 1, value=clean_val)
                
                # Auto-fit columns to prevent text clipping and close spacing
                for col in ws.columns:
                    max_len = 0
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                    
    # If no sheets were created, create at least one empty sheet
    if not wb.sheetnames:
        ws = wb.create_sheet(title="Sheet1")
        ws.cell(row=1, column=1, value="No tables or text found in PDF.")
        ws.views.sheetView[0].showGridLines = True
        
    # Save the optimized workbook
    wb.save(excel_path)
    print(f"Excel file created: {excel_path}")

def main():
    if len(sys.argv) < 3:
        print("Usage: python convert_pdf_to_excel.py <pdf_path> <excel_path>")
        sys.exit(1)
        
    pdf_path = sys.argv[1]
    excel_path = sys.argv[2]
    
    try:
        convert_pdf_to_excel(pdf_path, excel_path)
        print("Conversion completed successfully!")
    except Exception as e:
        print(f"Error during conversion: {str(e)}")
        sys.exit(1)

if __name__ == '__main__':
    main()
